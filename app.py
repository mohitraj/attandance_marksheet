from flask import Flask, render_template, request, send_from_directory, redirect, url_for, flash, session
import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from collections import defaultdict

app = Flask(__name__)
app.secret_key = 'your-secret-key-here-change-in-production'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['DOWNLOAD_FOLDER'] = 'downloads'
app.config['SAMPLE_FOLDER'] = 'static/samples'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['DOWNLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['SAMPLE_FOLDER'], exist_ok=True)


# ==================== LANDING PAGE ====================
@app.route('/')
def landing():
    return render_template('landing.html')


# ==================== PREPROCESS RAW ATTENDANCE FILE ====================
def preprocess_attendance_file(input_path, output_path):
    """
    Cleans the raw attendance Excel file into a flat single-header xlsx
    that pandas can read correctly with header=0.

    Source file structure:
      Rows 1-6  : metadata/title rows              -> DELETE
      Row 7     : session numbers (1, 2, 3 ...)    -> skip
      Row 8     : dates  e.g. '22-01'              -> used in column name
      Row 9     : # | Roll. No | Student Name | period numbers (3,4..) | Total | %age
      Row 10+   : student data

    Output (flat, single header row):
      Row 1  : Roll. No | Student Name | 22-01_3 | 22-01_4 | 29-01_3 ...
      Row 2+ : student data
    """
    wb_src = load_workbook(input_path)
    ws_src = wb_src.active

    roll_col = None
    name_col = None
    skip_cols = set()

    for c in range(1, ws_src.max_column + 1):
        val = ws_src.cell(9, c).value
        val_str = str(val).strip() if val is not None else ''
        if val_str == '#':
            skip_cols.add(c)
        elif val_str == 'Roll. No':
            roll_col = c
        elif val_str == 'Student Name':
            name_col = c
        elif val_str in ('Total', '%age'):
            skip_cols.add(c)

    keep_cols = [c for c in range(1, ws_src.max_column + 1) if c not in skip_cols]

    wb_new = Workbook()
    ws_new = wb_new.active

    # --- SINGLE flat header row (date_period format for internal merge use) ---
    for dest_c, src_c in enumerate(keep_cols, start=1):
        if src_c == roll_col:
            ws_new.cell(1, dest_c).value = 'Roll. No'
        elif src_c == name_col:
            ws_new.cell(1, dest_c).value = 'Student Name'
        else:
            date = ws_src.cell(8, src_c).value
            period = ws_src.cell(9, src_c).value
            ws_new.cell(1, dest_c).value = f"{date}_{period}"

    # --- DATA ROWS: source row 10 onwards -> destination row 2 onwards ---
    for src_row in range(10, ws_src.max_row + 1):
        dest_row = src_row - 8
        for dest_c, src_c in enumerate(keep_cols, start=1):
            ws_new.cell(dest_row, dest_c).value = ws_src.cell(src_row, src_c).value

    wb_new.save(output_path)


def build_display_header(ws, id_col_count=2):
    """
    Converts flat 'date_period'/'date_period_lab' header in row 1 into two rows:
      Row 1 (dates)  : date shown on EVERY attendance column including lab — no merging.
      Row 2 (periods): period number for every column, as-is.
    Inserts a new row 2, rewrites row 1 with dates.
    """
    max_col = ws.max_column
    flat_headers = [ws.cell(1, c).value for c in range(1, max_col + 1)]

    # Parse each column: (date, period, is_lab)
    parsed = []
    for h in flat_headers:
        if h in ('Roll. No', 'Student Name') or h is None:
            parsed.append((h, None, False))
        else:
            h_str = str(h)
            is_lab = h_str.endswith('_lab')
            base = h_str[:-4] if is_lab else h_str
            parts = base.split('_')
            if len(parts) == 2:
                parsed.append((parts[0], parts[1], is_lab))
            else:
                parsed.append((h_str, None, is_lab))

    # Insert new blank row 2 (student data shifts to row 3+)
    ws.insert_rows(2)

    # Clear row 1 before rewriting
    for c in range(1, max_col + 1):
        ws.cell(1, c).value = None

    # Write each column into rows 1 and 2
    for i, (date, period, is_lab) in enumerate(parsed):
        xl_col = i + 1
        if period is None:
            # Identity col (Roll. No / Student Name): label in row 1, blank in row 2
            ws.cell(1, xl_col).value = date
            ws.cell(2, xl_col).value = None
        else:
            # Every attendance col (lecture AND lab): date in row 1, period in row 2
            ws.cell(1, xl_col).value = date
            ws.cell(2, xl_col).value = period

    return ws


# ==================== ATTENDANCE MERGER ====================
def sort_key(col_name):
    is_lab = col_name.endswith('_lab')
    base = col_name[:-4] if is_lab else col_name
    parts = base.split('_')
    date_part = parts[0]
    period_part = parts[1] if len(parts) > 1 else '0'
    try:
        day, month = map(int, date_part.split('-'))
        period = int(period_part)
    except Exception:
        return (99, 99, 0, 99)
    # month, day → same date grouped; is_lab=0 (lecture) before is_lab=1 (lab); then period
    return (month, day, int(is_lab), period)


def setnumber(df, output_path):
    """
    Replace non-X attendance values with sequential lecture numbers per student row.
    Works on the flat single-header DataFrame (before display header is built).
    """
    data_cols = df.columns[2:]

    # Cast to object so pandas accepts mixed int/str values in the same column
    df = df.copy()
    df[data_cols] = df[data_cols].astype(object)

    def renumber_row(row):
        counter = 1
        for col in data_cols:
            val = str(row[col]).strip().upper()
            if val != 'X':
                row[col] = counter
                counter += 1
            else:
                row[col] = 'X'
        return row

    df[data_cols] = df[data_cols].apply(renumber_row, axis=1)
    df.to_excel(output_path, index=False)


def apply_styles_with_display_header(excel_path, styled_path):
    """
    1. Load the flat-header xlsx produced by setnumber().
    2. Build the two-row display header (date merged / period below).
    3. Apply all visual styles: Lecture row, Period row, X cells, borders, lab columns.
    """
    wb = load_workbook(excel_path)
    ws = wb.active

    # --- Step 1: Build two-row display header (inserts row 2) ---
    build_display_header(ws, id_col_count=2)

    # Now the sheet has:
    #   Row 1 : merged date groups  (e.g.  22-01  |  29-01  | ...)
    #   Row 2 : period numbers      (e.g.    3  4  |   3  4  | ...)
    #   Row 3+: student data

    total_cols = ws.max_column

    # --- Step 2: Insert Lecture row ABOVE the current row 1 ---
    ws.insert_rows(1)
    # Rows are now:
    #   Row 1 : (blank, to become Lecture row)
    #   Row 2 : merged dates
    #   Row 3 : period numbers
    #   Row 4+: student data

    # Fill Lecture row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    ws.cell(row=1, column=1).value = "Lecture"
    ws.cell(row=1, column=1).font = Font(size=14, bold=True)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')

    lecture_num = 1
    for col_idx in range(3, total_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = lecture_num
        lecture_num += 1
        cell.font = Font(size=14)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # --- Step 3: Style the date row (row 2) and period row (row 3) ---
    light_green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

    # Date row (row 2) — already has merged cells from build_display_header
    for col_idx in range(1, total_cols + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.fill = light_green_fill
        cell.font = Font(size=14, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Period Number row (row 3)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)
    ws.cell(row=3, column=1).value = "Period Number"
    ws.cell(row=3, column=1).font = Font(size=14, bold=True)
    ws.cell(row=3, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=3, column=1).fill = light_green_fill
    for col_idx in range(3, total_cols + 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.fill = light_green_fill
        cell.font = Font(size=14)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # --- Step 4: Style lab columns with yellow fill ---
    # Lab columns are identified from the original flat header which is now in row 2
    # But since we merged/rewrote row 2, detect lab cols from col index pattern.
    # We track via column letter: lab cols have '_lab' in the original flat headers.
    # Re-read original to identify them — they're stored as period=number in row 3.
    # Actually the simplest: check if the column was a lab column by examining the
    # flat-header file. We do this by re-reading the pre-styled file.
    wb_flat = load_workbook(excel_path)
    ws_flat = wb_flat.active
    flat_header = [ws_flat.cell(1, c).value for c in range(1, ws_flat.max_column + 1)]
    lab_col_indices = [i + 1 for i, h in enumerate(flat_header)
                       if isinstance(h, str) and h.endswith('_lab')]

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for col_idx in lab_col_indices:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            row[col_idx - 1].fill = yellow_fill

    # --- Step 5: Style X cells (red background, bold white) ---
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    for row in ws.iter_rows(min_row=4):
        for cell in row[2:]:
            if cell.value == 'X':
                cell.font = Font(bold=True, size=14, color="FFFFFF")
                cell.fill = red_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # --- Step 6: Borders and alignment for all cells ---
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                             min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            if isinstance(cell.value, (int, float)):
                cell.alignment = center

    wb.save(styled_path)


@app.route('/attendance')
def attendance():
    download_file = session.pop("download_file", None)
    return render_template('attendance.html', download_file=download_file)


@app.route('/attendance/upload', methods=['POST'])
def attendance_upload():
    try:
        lecture_file = request.files['lecture']
        lab_file = request.files['lab']
        if not lecture_file or not lab_file:
            flash("Both files are required.", "error")
            return redirect(url_for("attendance"))

        lecture_raw_path = os.path.join(app.config['UPLOAD_FOLDER'], "lecture_raw.xlsx")
        lab_raw_path = os.path.join(app.config['UPLOAD_FOLDER'], "lab_raw.xlsx")
        lecture_file.save(lecture_raw_path)
        lab_file.save(lab_raw_path)

        lecture_path = os.path.join(app.config['UPLOAD_FOLDER'], "lecture.xlsx")
        lab_path = os.path.join(app.config['UPLOAD_FOLDER'], "lab.xlsx")

        preprocess_attendance_file(lecture_raw_path, lecture_path)
        preprocess_attendance_file(lab_raw_path, lab_path)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = f"styled_final_{timestamp}.xlsx"
        final_path = os.path.join(app.config['DOWNLOAD_FOLDER'], f"final_{timestamp}.xlsx")
        styled_path = os.path.join(app.config['DOWNLOAD_FOLDER'], file_name)

        key_columns = ["Roll. No", "Student Name"]

        df1 = pd.read_excel(lecture_path, header=0)
        df2 = pd.read_excel(lab_path, header=0)

        # Cast key columns to string to prevent float/int mismatch causing cross-join
        for col in key_columns:
            if col in df1.columns:
                df1[col] = df1[col].astype(str).str.strip()
            if col in df2.columns:
                df2[col] = df2[col].astype(str).str.strip()

        # Drop rows where Roll. No is empty/nan (metadata artifact rows)
        df1 = df1[~df1["Roll. No"].isin(["", "nan", "None"])]
        df2 = df2[~df2["Roll. No"].isin(["", "nan", "None"])]

        # fillna only on attendance cols, not key cols (avoids corrupting Roll. No / Name)
        att_cols1 = [c for c in df1.columns if c not in key_columns]
        att_cols2 = [c for c in df2.columns if c not in key_columns]
        df1[att_cols1] = df1[att_cols1].fillna("A")
        df2[att_cols2] = df2[att_cols2].fillna("A")

        df2_renamed = df2.rename(
            columns={col: col + "_lab" for col in df2.columns if col not in key_columns}
        )

        merged_df = pd.merge(df1, df2_renamed, on=key_columns, how="inner")
        merged_df.columns = merged_df.columns.str.strip()
        date_cols = [col for col in merged_df.columns if col not in key_columns]

        sort_cols = sorted(date_cols, key=sort_key)
        final_df = merged_df.loc[:, key_columns + sort_cols]

        # setnumber writes flat single-header xlsx (unchanged from before)
        setnumber(final_df, final_path)

        # NEW: apply_styles_with_display_header builds two-row header + all styling
        apply_styles_with_display_header(final_path, styled_path)

        session["download_file"] = file_name
        flash("File processed successfully! Thanks to Mohit", "success")
        return redirect(url_for("attendance"))

    except Exception as e:
        flash(f"Error: {str(e)}", "error")
        return redirect(url_for("attendance"))


@app.route('/attendance/download/<filename>')
def attendance_download(filename):
    return send_from_directory(app.config['DOWNLOAD_FOLDER'], filename, as_attachment=True)


# ==================== EXCEL JOINER ====================
@app.route('/joiner')
def joiner():
    return render_template('joiner.html')


if __name__ == '__main__':
    app.run(debug=True, host="0.0.0.0")
